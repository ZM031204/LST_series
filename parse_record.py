import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill,Font
from dotenv import load_dotenv
import ee

def date_line(year):
    month = list(range(1, 13))
    return [year] + month

def parse_record(file_path, start_year=1985, end_year=2024):
    file_dir = os.path.dirname(file_path)
    df = pd.read_csv(file_path)
    df = df.sort_values(by=['city', 'year', 'month'])
    city_list = df['city'].unique()

    total_year = range(start_year, end_year+1)
    cloud_high_color = Font(color='FF0000')
    cloud_median_color = Font(color='0000FF')
    cloud_low_color = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid') # green for low cloud
    cover_high_color = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid') # green for high cover
    cover_low_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # yellow for low cover
    note_city = Workbook()
    note_city.remove(note_city['Sheet'])
    for city in city_list:
        this_city_data = df[df['city'] == city]
        file_name = f"records/{city}.csv"
        note_city_ws = note_city.create_sheet(title=city)
        current_row = 1  # 添加行追踪器
        os.makedirs(os.path.dirname(file_name), exist_ok=True)
        print(f'executing {city}')
        with open(file_name, "w", newline='', encoding='utf-8') as f:
            for index, year in enumerate(total_year):
                print(f'executing {year}')
                f.write(','.join(map(str, date_line(year))) + '\n')
                note_city_ws.append(date_line(year))
                current_year_data = this_city_data[this_city_data['year'] == year]
                property_list = ['toa_image_porpotion','sr_image_porpotion','toa_cloud_ratio','sr_cloud_ratio']
                for pid, pro in enumerate(property_list):
                    value_list = ['/'] * 12
                    for index, row in current_year_data.iterrows():
                        value_list[row['month']-1] = row[pro]
                    row_line = [pro] + value_list
                    note_city_ws.append(row_line)
                    # change color of cloud ratio
                    if pro in ['toa_cloud_ratio', 'sr_cloud_ratio']:
                        for i in range(1, 13):
                            if row_line[i] != '/':
                                if float(row_line[i]) > 10:
                                    note_city_ws.cell(column=i+1, row=current_row+pid+1).font = cloud_high_color
                                elif float(row_line[i]) < 5:
                                    note_city_ws.cell(column=i+1, row=current_row+pid+1).fill = cloud_low_color
                                else:
                                    note_city_ws.cell(column=i+1, row=current_row+pid+1).font = cloud_median_color
                    elif pro in ['toa_image_porpotion', 'sr_image_porpotion']:
                        for i in range(1, 13):
                            if row_line[i] != '/':
                                if float(row_line[i]) < 0.9:
                                    note_city_ws.cell(column=i+1, row=current_row+pid+1).fill = cover_low_color
                                else:
                                    note_city_ws.cell(column=i+1, row=current_row+pid+1).fill = cover_high_color

                    f.write(','.join(map(str, row_line)) + '\n')
                current_row += 5  # 每年的数据占用5行（1行年份+4行属性）
    note_city.save(os.path.join(file_dir, 'city_quality_records.xlsx'))

def get_geo_boundary():
    load_dotenv()
    project_name = os.getenv('PROJECT_NAME')
    ee.Initialize(project=project_name)

    asset_path = 'projects/ee-channingtong/assets/'
    total_boundary = ee.FeatureCollection(asset_path + 'YZBboundary')
    geo_boundary_dict = {}
    for city_boundary in total_boundary.getInfo()['features']:
        city_name = city_boundary['properties']['市名']
        city_outbound = ee.Geometry(city_boundary['geometry']).bounds().getInfo()
        geo_boundary_dict[city_name] = city_outbound

    return geo_boundary_dict

def reverse_parse_record(tag_file_path, raw_file_path, start_year=1985, end_year=2024):
    workbook = load_workbook(tag_file_path)
    year_list = list(range(start_year, end_year+1))
    month_list = list(range(1, 13))
    valid_records = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for year in year_list:
            row = 1 + (year - start_year) * 5
            for month in month_list:
                cell = sheet.cell(row=row, column=month+1)
                if cell.fill.start_color.index != "00000000":
                    valid_records.append((sheet_name,year, month))

    raw_data = pd.read_csv(raw_file_path)
    geo_boundary_dict = get_geo_boundary()

    output_df = pd.DataFrame(columns=['city', 'year', 'month', 'day', 'geometry'])
    for record in valid_records:
        str_string = f"{record[0]},{record[1]},{record[2]}"
        for row in raw_data.iterrows():
            row_data = row[1]
            identify = f"{row_data['city']},{row_data['year']},{row_data['month']}"
            if str_string == identify:
                output_df.loc[len(output_df)] = {
                    'city': row_data['city'],
                    'year': row_data['year'],
                    'month': row_data['month'],
                    'day': row_data['day'],
                    'geometry': geo_boundary_dict[row_data['city']]
                }
                print(f"added {record[0]},{record[1]},{record[2]}")
                break
    # output as json file using pandas
    output_df.to_json('remote_sensing_record.json', orient='records', lines=True, force_ascii=False)
    return True